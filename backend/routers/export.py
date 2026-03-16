from fastapi import APIRouter, Depends, HTTPException, Query
from routers.auth import require_analyst
from fastapi.responses import StreamingResponse
from database import execute_query
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import re
import threading
import uuid
from datetime import datetime, timedelta

router = APIRouter(prefix="/export", tags=["Export"])

# Job store em memória — jobs expiram após 10 minutos
_export_jobs: dict = {}
_jobs_lock = threading.Lock()

def _cleanup_expired_jobs():
    cutoff = datetime.utcnow() - timedelta(minutes=10)
    with _jobs_lock:
        expired = [jid for jid, j in _export_jobs.items() if j["created_at"] < cutoff]
        for jid in expired:
            del _export_jobs[jid]


def _generate_excel_job(job_id, table_name, date_column, date_from, date_to):
    """Gera Excel em background thread, atualizando progresso."""
    try:
        with _jobs_lock:
            _export_jobs[job_id]["status"] = "processing"

        sql = f'SELECT * FROM public."{table_name}"'
        params = []
        if date_column and (date_from or date_to):
            conditions = []
            if date_from:
                conditions.append(f'"{date_column}" >= %s')
                params.append(date_from)
            if date_to:
                conditions.append(f'"{date_column}" <= %s')
                params.append(date_to + ' 23:59:59')
            sql += ' WHERE ' + ' AND '.join(conditions)
        sql += ' LIMIT 500000'

        rows = execute_query(sql, tuple(params) if params else None)
        total = len(rows)

        if not rows:
            with _jobs_lock:
                _export_jobs[job_id]["status"] = "error"
                _export_jobs[job_id]["error"] = "Nenhum dado encontrado"
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = table_name[:31]
        columns = list(rows[0].keys())

        header_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="00D4FF", size=11)
        header_border = Border(bottom=Side(style='medium', color="00D4FF"))
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=False)

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = header_border
            cell.alignment = header_align

        row_fill_even = PatternFill(start_color="111720", end_color="111720", fill_type="solid")
        row_fill_odd = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
        row_font = Font(name="Calibri", color="E2E8F0", size=10)
        row_align = Alignment(horizontal='left', vertical='center')

        BATCH_SIZE = 1000
        for row_idx, row in enumerate(rows, start=2):
            fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
            for col_idx, col_name in enumerate(columns, start=1):
                value = row[col_name]
                if hasattr(value, 'isoformat'):
                    value = value.strftime('%d/%m/%Y %H:%M') if hasattr(value, 'hour') else value.strftime('%d/%m/%Y')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = row_font
                cell.fill = fill
                cell.alignment = row_align

            rows_done = row_idx - 1
            if rows_done % BATCH_SIZE == 0 or rows_done == total:
                progress = int(100 * rows_done / total) if total > 0 else 100
                with _jobs_lock:
                    if job_id in _export_jobs:
                        _export_jobs[job_id]["rows_done"] = rows_done
                        _export_jobs[job_id]["progress"] = progress

        for col_idx, col_name in enumerate(columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(col_name))
            for sample_row in range(2, min(total + 2, 52)):
                cell_val = ws.cell(row=sample_row, column=col_idx).value
                if cell_val:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        ws.freeze_panes = "A2"
        wb.properties.title = f"Hinova DW — {table_name}"
        wb.properties.creator = "Hinova DW Monitor"

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        date_suffix = ""
        if date_from and date_to:
            date_suffix = f"_{date_from}_a_{date_to}"
        elif date_from:
            date_suffix = f"_a_partir_{date_from}"
        elif date_to:
            date_suffix = f"_ate_{date_to}"
        filename = f"hinova_dw_{table_name}{date_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        with _jobs_lock:
            if job_id in _export_jobs:
                _export_jobs[job_id]["status"] = "done"
                _export_jobs[job_id]["progress"] = 100
                _export_jobs[job_id]["rows_done"] = total
                _export_jobs[job_id]["result_buffer"] = buffer
                _export_jobs[job_id]["filename"] = filename

    except Exception as e:
        with _jobs_lock:
            if job_id in _export_jobs:
                _export_jobs[job_id]["status"] = "error"
                _export_jobs[job_id]["error"] = str(e)


@router.get("/tables")
def list_exportable_tables(token_data: dict = Depends(require_analyst)):
    """Lista tabelas com rowcount e colunas de data detectadas."""
    rows = execute_query("""
        SELECT
            t.relname AS table_name,
            t.n_live_tup AS rows_estimate,
            pg_size_pretty(pg_total_relation_size(t.relid)) AS size_pretty,
            COALESCE(
                (SELECT string_agg(column_name, ',' ORDER BY ordinal_position)
                 FROM information_schema.columns c
                 WHERE c.table_name = t.relname
                   AND c.table_schema = 'public'
                   AND c.data_type IN (
                       'date', 'timestamp without time zone',
                       'timestamp with time zone'
                   )
                ),
                ''
            ) AS date_columns
        FROM pg_stat_user_tables t
        WHERE t.schemaname = 'public'
        ORDER BY pg_total_relation_size(t.relid) DESC
    """)

    return {
        "tables": [
            {
                "name": r["table_name"],
                "rows_estimate": r["rows_estimate"],
                "size_pretty": r["size_pretty"],
                "date_columns": r["date_columns"].split(",") if r["date_columns"] else []
            }
            for r in rows
        ],
        "updated_at": datetime.utcnow().isoformat()
    }


@router.get("/preview/{table_name}")
def preview_table(table_name: str, token_data: dict = Depends(require_analyst)):
    """Retorna colunas e primeiras 5 linhas da tabela."""
    if not re.match(r'^[a-zA-Z0-9_]+$', table_name):
        raise HTTPException(status_code=400, detail="Nome de tabela inválido")

    exists = execute_query(
        "SELECT 1 FROM information_schema.tables "
        "WHERE table_schema = 'public' AND table_name = %s",
        (table_name,)
    )
    if not exists:
        raise HTTPException(status_code=404, detail=f"Tabela '{table_name}' não encontrada")

    rows = execute_query(
        f'SELECT * FROM public."{table_name}" LIMIT 5'
    )

    columns = list(rows[0].keys()) if rows else []

    return {
        "table": table_name,
        "columns": columns,
        "preview": rows,
        "updated_at": datetime.utcnow().isoformat()
    }


@router.get("/download/{table_name}")
def export_table(
    table_name: str,
    date_column: str = Query(default=None),
    date_from: str = Query(default=None),
    date_to: str = Query(default=None),
    limit: int = Query(default=500000),
    token_data: dict = Depends(require_analyst),
):
    """Exporta tabela completa ou filtrada por período para Excel."""
    if not re.match(r'^[a-zA-Z0-9_]+$', table_name):
        raise HTTPException(status_code=400, detail="Nome de tabela inválido")

    if date_column and not re.match(r'^[a-zA-Z0-9_]+$', date_column):
        raise HTTPException(status_code=400, detail="Nome de coluna inválido")

    # Enforce max limit
    limit = min(int(limit), 500000)

    sql = f'SELECT * FROM public."{table_name}"'
    params = []

    if date_column and (date_from or date_to):
        conditions = []
        if date_from:
            conditions.append(f'"{date_column}" >= %s')
            params.append(date_from)
        if date_to:
            conditions.append(f'"{date_column}" <= %s')
            params.append(date_to + ' 23:59:59')
        sql += ' WHERE ' + ' AND '.join(conditions)

    sql += f' LIMIT {limit}'

    rows = execute_query(sql, tuple(params) if params else None)

    if not rows:
        raise HTTPException(status_code=404, detail="Nenhum dado encontrado para os filtros informados")

    # Generate Excel in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = table_name[:31]

    columns = list(rows[0].keys())

    # Header style
    header_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="00D4FF", size=11)
    header_border = Border(bottom=Side(style='medium', color="00D4FF"))
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=False)

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_align

    # Row styles
    row_fill_even = PatternFill(start_color="111720", end_color="111720", fill_type="solid")
    row_fill_odd = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
    row_font = Font(name="Calibri", color="E2E8F0", size=10)
    row_align = Alignment(horizontal='left', vertical='center')

    for row_idx, row in enumerate(rows, start=2):
        fill = row_fill_even if row_idx % 2 == 0 else row_fill_odd
        for col_idx, col_name in enumerate(columns, start=1):
            value = row[col_name]
            if hasattr(value, 'isoformat'):
                value = value.strftime('%d/%m/%Y %H:%M') if hasattr(value, 'hour') else value.strftime('%d/%m/%Y')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = row_font
            cell.fill = fill
            cell.alignment = row_align

    # Auto-width columns
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for row_idx in range(2, min(len(rows) + 2, 52)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    # Freeze header
    ws.freeze_panes = "A2"

    # Metadata
    wb.properties.title = f"Hinova DW — {table_name}"
    wb.properties.creator = "Hinova DW Monitor"

    # Stream
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    date_suffix = ""
    if date_from and date_to:
        date_suffix = f"_{date_from}_a_{date_to}"
    elif date_from:
        date_suffix = f"_a_partir_{date_from}"
    elif date_to:
        date_suffix = f"_ate_{date_to}"

    filename = f"hinova_dw_{table_name}{date_suffix}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@router.post("/start/{table_name}")
def start_export(
    table_name: str,
    date_column: str = Query(default=None),
    date_from: str = Query(default=None),
    date_to: str = Query(default=None),
    token_data: dict = Depends(require_analyst),
):
    """Inicia exportação assíncrona e retorna job_id."""
    if not re.match(r'^[a-zA-Z0-9_]+$', table_name):
        raise HTTPException(status_code=400, detail="Nome de tabela inválido")
    if date_column and not re.match(r'^[a-zA-Z0-9_]+$', date_column):
        raise HTTPException(status_code=400, detail="Nome de coluna inválido")

    exists = execute_query(
        "SELECT 1 FROM information_schema.tables "
        "WHERE table_schema = 'public' AND table_name = %s",
        (table_name,)
    )
    if not exists:
        raise HTTPException(status_code=404, detail=f"Tabela '{table_name}' não encontrada")

    count_sql = f'SELECT COUNT(*) AS total FROM public."{table_name}"'
    count_params = []
    if date_column and (date_from or date_to):
        conditions = []
        if date_from:
            conditions.append(f'"{date_column}" >= %s')
            count_params.append(date_from)
        if date_to:
            conditions.append(f'"{date_column}" <= %s')
            count_params.append(date_to + ' 23:59:59')
        count_sql += ' WHERE ' + ' AND '.join(conditions)

    count_result = execute_query(count_sql, tuple(count_params) if count_params else None)
    total_rows = min(count_result[0]["total"], 500000) if count_result else 0

    job_id = str(uuid.uuid4())[:8]
    with _jobs_lock:
        _export_jobs[job_id] = {
            "status": "pending",
            "progress": 0,
            "rows_done": 0,
            "total_rows": total_rows,
            "table_name": table_name,
            "error": None,
            "result_buffer": None,
            "filename": None,
            "created_at": datetime.utcnow()
        }

    thread = threading.Thread(
        target=_generate_excel_job,
        args=(job_id, table_name, date_column, date_from, date_to),
        daemon=True
    )
    thread.start()

    return {
        "job_id": job_id,
        "total_rows": total_rows,
        "table_name": table_name,
        "started_at": datetime.utcnow().isoformat()
    }


@router.get("/progress/{job_id}")
def get_export_progress(job_id: str, token_data: dict = Depends(require_analyst)):
    """Retorna progresso atual do job de exportação."""
    _cleanup_expired_jobs()

    with _jobs_lock:
        job = _export_jobs.get(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado ou expirado")

    return {
        "job_id": job_id,
        "status": job["status"],
        "progress": job["progress"],
        "rows_done": job["rows_done"],
        "total_rows": job["total_rows"],
        "table_name": job["table_name"],
        "error": job["error"],
        "filename": job["filename"],
        "checked_at": datetime.utcnow().isoformat()
    }


@router.get("/result/{job_id}")
def download_export_result(job_id: str, token_data: dict = Depends(require_analyst)):
    """Baixa o arquivo Excel gerado pelo job."""
    with _jobs_lock:
        job = _export_jobs.get(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job não encontrado ou expirado")
    if job["status"] != "done":
        raise HTTPException(status_code=400, detail=f"Job ainda não concluído. Status: {job['status']}")
    if not job["result_buffer"]:
        raise HTTPException(status_code=500, detail="Buffer do arquivo não encontrado")

    job["result_buffer"].seek(0)

    def _remove_job():
        import time
        time.sleep(30)
        with _jobs_lock:
            _export_jobs.pop(job_id, None)
    threading.Thread(target=_remove_job, daemon=True).start()

    return StreamingResponse(
        job["result_buffer"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{job["filename"]}"'}
    )
